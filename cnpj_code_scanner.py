#!/usr/bin/env python3
"""Levantamento de impacto para adequação ao CNPJ alfanumérico."""

from __future__ import annotations

import argparse
import hashlib
import html
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence
from xml.sax.saxutils import escape as xml_escape
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_DISPONIVEL = True
except Exception:
    OPENPYXL_DISPONIVEL = False

EXTENSOES_SUPORTADAS = {
    ".cs", ".vb", ".fs", ".fsx", ".java", ".kt", ".kts", ".scala", ".go", ".rs", ".php", ".py", ".rb",
    ".js", ".jsx", ".ts", ".tsx", ".mjs", ".cjs", ".vue", ".svelte", ".html", ".htm", ".cshtml", ".razor",
    ".aspx", ".ascx", ".css", ".scss", ".sass", ".less", ".sql", ".ddl", ".dml", ".psql", ".hql", ".ktr", ".kjb",
    ".json", ".xml", ".yml", ".yaml", ".config", ".ini", ".properties", ".env", ".csproj", ".vbproj", ".gradle",
    ".tf", ".tfvars", ".ps1", ".bat", ".cmd", ".sh", ".bash",
}

DIRETORIOS_EXCLUIDOS_PADRAO = {
    ".git", ".svn", ".hg", ".idea", ".vscode", "node_modules", "dist", "build", "target", "bin", "obj",
    "vendor", "packages", "coverage", "wwwroot", "publicacoes", "__pycache__", ".venv", "venv",
}

MARCADORES_PROJETO = {
    ".git", ".sln", "pom.xml", "build.gradle", "settings.gradle", "settings.gradle.kts", "package.json",
    "angular.json", "composer.json", "Gemfile", "pyproject.toml", "setup.py", "requirements.txt", "Cargo.toml",
}

NOMES_PIPELINE_INFRA = {
    ".gitlab-ci.yml", "azure-pipelines.yml", "jenkinsfile", "docker-compose.yml", "dockerfile",
}
PADROES_NOME_PIPELINE_INFRA = (
    re.compile(r"(?i)^docker-compose\..+\.ya?ml$"),
)
TRECHOS_PIPELINE_INFRA = (
    "/.github/workflows/", "/.gitlab/", "/jenkins/", "/helm/", "/k8s/", "/kubernetes/", "/manifests/", "/deploy/", "/argocd/",
)

RE_NOME_IGNORADO = re.compile(r"\.(?:min\.(?:js|css)|map|lock)$", re.IGNORECASE)
NOMES_IGNORADOS = {"package-lock.json", "yarn.lock", "pnpm-lock.yaml", "composer.lock"}
RE_ANCORA_CNPJ = re.compile(r"(?i)\b(cnpj|cpf\s*/\s*cnpj|cpfcnpj|cnpjcpf|nr_cnpj|num_cnpj|cd_cnpj|cnpj_nr|cnpj_num)\b")
RE_MASCARA_CLASSICA = re.compile(r"00\.000\.000/0000-00|99\.999\.999/9999-99|##\.###\.###/####-##")
RE_ALVO_NAO_CNPJ = re.compile(
    r"(?i)\b(telefone|celular|email|e-?mail|fax|cep|endereco|endereço|inscricaoestadual|inscricao_estadual|inscrição estadual|inscricaomunicipal|inscricao_municipal|inscrição municipal)\b"
)
RE_ALVO_CNPJ = re.compile(
    r"(?i)\b(cnpj|cpfcnpj|cpf_cnpj|cpf\s*/\s*cnpj|documento(pj)?|pessoa_juridica|pessoajuridica|cnpj_matriz)\b"
)
RE_METADATA_PIPELINE = re.compile(
    r"(?i)\b(application_name|app_name|container_name|image|job|stage|service|deploy|environment|env|connection_string)\b"
)
RE_EVIDENCIA_FORTE_PIPELINE = re.compile(
    r"(?i)\\d\{14\}|\[0-9\]\{14\}|\b\d{14}\b|isdigit|isnumeric|regex|replace|re\.sub|max_length\s*=\s*14|"
    r"varchar\s*\(\s*14\s*\)|char\s*\(\s*14\s*\)|\bcnpj\b.{0,50}(dto|schema|payload|contract|endpoint|api)"
)

COLUNAS_SAIDA = [
    "projeto",
    "arquivo",
    "arquivo_absoluto",
    "extensao",
    "camada",
    "linha_inicial",
    "padrao_identificado",
    "categoria_tecnica",
    "trecho",
    "contexto",
    "pontuacao",
    "origem_artefato",
    "artefato_gerado",
    "artefato_terceiro",
    "evidencia_contextual",
    "prioridade_tratamento",
    "identificador_consolidacao",
    "natureza_intervencao",
    "classificacao_resultado",
]


@dataclass(frozen=True)
class PadraoDeteccao:
    padrao_identificado: str
    expressao: re.Pattern[str]
    categoria_tecnica: str
    prioridade_tratamento: str
    pontuacao: int
    natureza_intervencao: str
    classificacao_resultado: str
    exige_contexto: bool = True


@dataclass
class OcorrenciaImpacto:
    projeto: str
    arquivo: str
    arquivo_absoluto: str
    extensao: str
    camada: str
    linha_inicial: int
    padrao_identificado: str
    categoria_tecnica: str
    trecho: str
    contexto: str
    pontuacao: int
    origem_artefato: str
    artefato_gerado: str
    artefato_terceiro: str
    evidencia_contextual: str
    prioridade_tratamento: str
    identificador_consolidacao: str
    natureza_intervencao: str
    classificacao_resultado: str

    def para_linha(self) -> list[str | int]:
        return [getattr(self, coluna) for coluna in COLUNAS_SAIDA]


PADROES_DETECCAO: tuple[PadraoDeteccao, ...] = (
    PadraoDeteccao("Regex de 14 dígitos", re.compile(r"\\d\{14\}|\[0-9\]\{14\}|\b\d{14}\b"), "Validação", "critica", 100, "validacao_numerica_restritiva", "adequacao_obrigatoria", True),
    PadraoDeteccao("Regex CNPJ formatado legado", re.compile(r"\\d\{2\}\\\.\\d\{3\}\\\.\\d\{3\}[/\\/]\\d\{4\}-\\d\{2\}|\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}"), "Validação", "critica", 98, "validacao_estrutural_legada", "adequacao_obrigatoria", True),
    PadraoDeteccao("Validação por método CNPJ", re.compile(r"(?i)\bvalid\w*cnpj\b|\bcnpj\w*valid\b|\bisValidCnpj\b|\bcnpjValido\b|\bcheckCnpj\b|\bvalidaCnpj\b"), "Validação", "critica", 97, "validacao_estrutural_legada", "adequacao_obrigatoria", False),
    PadraoDeteccao("Validação numérica estrita", re.compile(r"(?i)\b(isdigit|isnumeric|numeric)\b"), "Validação", "alta", 93, "validacao_numerica_restritiva", "adequacao_obrigatoria", True),
    PadraoDeteccao("Sanitização para apenas dígitos", re.compile(r"(?i)regex\.replace\(.+\^\\d|\.replace\s*\(/\\D/g|re\.sub\(.+\\D|onlynumbers|apenasnumeros|somentenumeros"), "Validação", "alta", 92, "validacao_numerica_restritiva", "adequacao_obrigatoria", True),
    PadraoDeteccao("Comprimento fixo igual a 14", re.compile(r"(?i)\blen\s*\(?\s*\w+\s*\)?\s*==\s*14\b|\.length\s*==\s*14\b|max_length\s*=\s*14\b"), "Validação", "alta", 91, "comprimento_fixo_legado", "adequacao_obrigatoria", True),
    PadraoDeteccao("Máscara clássica de CNPJ", re.compile(r"00\.000\.000/0000-00|99\.999\.999/9999-99|##\.###\.###/####-##|__\.___.___/____-__"), "Interface", "alta", 90, "mascara_legada", "adequacao_obrigatoria", True),
    PadraoDeteccao("Algoritmo legado de dígito verificador", re.compile(r"(?i)d[ií]gito\s+verificador|m[oó]dulo\s*11|calcula[_-]?digito|primeiro[_-]?digito|segundo[_-]?digito|5\s*,\s*4\s*,\s*3\s*,\s*2\s*,\s*9\s*,\s*8"), "Validação", "alta", 89, "validacao_estrutural_legada", "adequacao_obrigatoria", True),
    PadraoDeteccao("Restrição de persistência CNPJ", re.compile(r"(?i)\b(varchar|nvarchar|char)\s*\(\s*14\s*\)|\b(nr_cnpj|num_cnpj|cd_cnpj|cnpj_nr|cnpj_num|ds_cnpj|tx_cnpj)\b|\b(create\s+(unique\s+)?index|constraint)\b[^\n]{0,100}\bcnpj\b"), "Persistência", "alta", 88, "restricao_persistencia", "adequacao_obrigatoria", True),
    PadraoDeteccao("Contrato de integração com CNPJ", re.compile(r"(?i)\b(dto|schema|request|response|payload|contract|endpoint|api|route|url|uri|path)\b[^\n]{0,80}\bcnpj\b|\bcnpj\b[^\n]{0,80}\b(dto|schema|request|response|payload|contract|endpoint|api)\b"), "Integração", "media", 72, "contrato_integracao", "adequacao_potencial", True),
    PadraoDeteccao("Referência de interface CNPJ", re.compile(r"(?i)\b(label|placeholder|title|hint|aria-label|formcontrolname|ng-model|id|name|mensagem|erro|msg)\b[^\n]{0,100}\bcnpj\b"), "Interface", "baixa", 55, "referencia_interface", "evidencia_contextual", False),
)


def argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Levantamento de impacto para adequação ao CNPJ alfanumérico.")
    parser.add_argument("raiz", nargs="?", default=".", help="Diretório raiz de análise.")
    parser.add_argument("--diretorio-saida", default="./resultado_cnpj_scan", help="Diretório de saída.")
    parser.add_argument("--encoding", default="utf-8", help="Encoding preferencial.")
    parser.add_argument("--tamanho-maximo-kb", type=int, default=1024, help="Tamanho máximo de arquivo (KB).")
    parser.add_argument("--extensoes-adicionais", nargs="*", help="Extensões extras para varredura.")
    parser.add_argument("--diretorios-excluidos", nargs="*", help="Nomes de diretórios excluídos.")
    parser.add_argument("--modo-agrupamento-projeto", choices=("auto", "topdir", "none"), default="auto")
    parser.add_argument("--janela-contexto", type=int, default=4, help="Janela contextual (±N linhas).")
    parser.add_argument("--desabilitar-html", action="store_true", help="Desabilita geração de HTML.")
    return parser.parse_args()


def detectar_projeto(caminho: Path, raiz: Path, modo: str) -> str:
    rel = caminho.relative_to(raiz)
    if modo == "none":
        return "SEM_GRUPO"
    if modo == "topdir":
        return rel.parts[0] if len(rel.parts) > 1 else "RAIZ"
    atual = caminho.parent
    while True:
        if any((atual / marcador).exists() for marcador in MARCADORES_PROJETO):
            return "RAIZ" if atual == raiz else atual.name
        if atual == raiz:
            break
        atual = atual.parent
    return rel.parts[0] if len(rel.parts) > 1 else "RAIZ"


def inferir_camada(caminho: Path) -> str:
    ext = caminho.suffix.lower()
    if ext in {".sql", ".ddl", ".dml", ".psql", ".hql", ".ktr", ".kjb"}:
        return "Persistência"
    if ext in {".js", ".ts", ".jsx", ".tsx", ".vue", ".svelte", ".html", ".htm", ".css", ".scss", ".sass", ".less", ".cshtml", ".razor", ".aspx", ".ascx"}:
        return "Interface"
    if ext in {".json", ".xml", ".config", ".yml", ".yaml", ".ini", ".properties", ".env", ".csproj", ".vbproj", ".gradle", ".tf", ".tfvars"}:
        return "Configuração"
    if ext in {".ps1", ".bat", ".cmd", ".sh", ".bash"}:
        return "Automação"
    return "Aplicação"


def classificar_origem_artefato(caminho: Path) -> str:
    p = caminho.as_posix().lower()
    nome = caminho.name.lower()
    if any(token in p for token in ("/vendor/", "/third_party/", "/third-party/", "/plugins/", "/plugin/", "/lib/", "/node_modules/", "/inputmask/")):
        return "biblioteca_terceiros"
    if "migration" in p or "migrations" in p:
        return "migracao"
    if "snapshot" in p or nome.endswith("snapshot.cs"):
        return "snapshot_modelo"
    if "designer" in p or nome.endswith(".designer.cs"):
        return "arquivo_designer"
    if any(token in p for token in ("/generated/", "/codegen/", "/gen/")) or "autogenerated" in nome or nome.endswith(".g.cs"):
        return "codigo_gerado"
    return "codigo_autoral"


def arquivo_pipeline_infra(caminho: Path) -> bool:
    p = caminho.as_posix().lower()
    nome = caminho.name.lower()
    if nome in NOMES_PIPELINE_INFRA:
        return True
    if any(expr.search(nome) for expr in PADROES_NOME_PIPELINE_INFRA):
        return True
    if any(token in p for token in TRECHOS_PIPELINE_INFRA):
        return True
    if nome.endswith((".yaml", ".yml", ".env")) and any(token in nome for token in ("pipeline", "compose", "deploy", "helm", "k8s", "kubernetes")):
        return True
    return False


def iterar_arquivos(
    raiz: Path,
    extensoes: set[str],
    diretorios_excluidos: set[str],
    tamanho_maximo_kb: int,
    diretorio_saida: Path,
) -> Iterable[Path]:
    maximo_bytes = tamanho_maximo_kb * 1024
    diretorio_saida = diretorio_saida.resolve()

    for caminho in raiz.rglob("*"):
        if caminho.is_dir():
            continue
        try:
            resolvido = caminho.resolve()
            if resolvido == diretorio_saida or diretorio_saida in resolvido.parents:
                continue
        except OSError:
            continue
        if caminho.suffix.lower() not in extensoes:
            continue
        if any(parte.lower() in diretorios_excluidos for parte in caminho.parts):
            continue
        if RE_NOME_IGNORADO.search(caminho.name) or caminho.name.lower() in NOMES_IGNORADOS:
            continue
        try:
            if caminho.stat().st_size > maximo_bytes:
                continue
        except OSError:
            continue
        yield caminho


def ler_linhas(caminho: Path, encoding_preferencial: str) -> Sequence[str]:
    for enc in (encoding_preferencial, "utf-8", "latin-1"):
        try:
            return caminho.read_text(encoding=enc).splitlines()
        except Exception:
            pass
    return []


def extrair_alvos_semanticos(texto: str) -> str:
    expressoes = [
        r'(?i)\b(id|name|for|placeholder|ng-model|formControlName)\s*=\s*["\']([^"\']+)["\']',
        r'(?i)\b(const|let|var|public static|private static)\s+([A-Za-z_][\w$]*)',
        r'(?i)\b([A-Za-z_][\w$]*)\s*:\s*\{',
        r'(?i)\b([A-Za-z_][\w$]*)\s*=',
    ]
    alvos: list[str] = []
    for expr in expressoes:
        for encontrado in re.finditer(expr, texto):
            if encontrado.lastindex and encontrado.lastindex >= 2:
                alvos.append((encontrado.group(2) or "").strip())
            elif encontrado.lastindex and encontrado.lastindex >= 1:
                alvos.append((encontrado.group(1) or "").strip())
    return " ".join(alvos).lower()


def alvo_semantico_valido(contexto: str, linha: str) -> bool:
    alvo = extrair_alvos_semanticos(contexto + "\n" + linha)
    if alvo and RE_ALVO_NAO_CNPJ.search(alvo) and not RE_ALVO_CNPJ.search(alvo):
        return False
    if RE_ALVO_NAO_CNPJ.search(linha) and not RE_ALVO_CNPJ.search(linha):
        return False
    return True


def validar_contexto(padrao: PadraoDeteccao, linhas: Sequence[str], indice: int, janela: int) -> bool:
    inicio = max(0, indice - janela)
    fim = min(len(linhas), indice + janela + 1)
    bloco = "\n".join(linhas[inicio:fim])
    linha = linhas[indice]

    if padrao.exige_contexto and not (RE_ANCORA_CNPJ.search(bloco) or RE_MASCARA_CLASSICA.search(bloco)):
        return False
    if not alvo_semantico_valido(bloco, linha):
        return False

    if "dígito verificador" in padrao.padrao_identificado.lower():
        return bool(re.search(r"(?i)valida|pesos|\b14\b", bloco) or re.search(r"5\s*,\s*4\s*,\s*3\s*,\s*2", bloco))

    if "máscara" in padrao.padrao_identificado.lower() and re.search(r"(?i)\$\.fn\.inputmask|data-inputmask|\bmaskset\b|\balias(es)?\b|\binputmask\(fn\)", linha):
        return False

    return True


def gerar_identificador_consolidacao(
    arquivo: str,
    regiao: int,
    natureza_intervencao: str,
    classificacao_resultado: str,
    padrao_identificado: str,
) -> str:
    base = f"{arquivo}|{regiao}|{natureza_intervencao}|{classificacao_resultado}|{padrao_identificado}"
    return hashlib.sha1(base.encode()).hexdigest()[:12]


def varrer_arquivo(caminho: Path, raiz: Path, args: argparse.Namespace) -> list[OcorrenciaImpacto]:
    linhas = ler_linhas(caminho, args.encoding)
    if not linhas:
        return []

    arquivo_rel = caminho.relative_to(raiz).as_posix()
    origem = classificar_origem_artefato(caminho)
    eh_pipeline = arquivo_pipeline_infra(caminho)
    eh_config = caminho.suffix.lower() in {".yaml", ".yml", ".env", ".ini", ".config", ".properties", ".json", ".xml"}

    ocorrencias: list[OcorrenciaImpacto] = []
    vistos: set[tuple[int, str, str, str, str]] = set()

    for indice, linha in enumerate(linhas):
        for padrao in PADROES_DETECCAO:
            if not padrao.expressao.search(linha):
                continue

            contexto_ok = validar_contexto(padrao, linhas, indice, args.janela_contexto)
            classificacao = padrao.classificacao_resultado if contexto_ok else "ocorrencia_descartada"
            bloco_contexto = "\n".join(linhas[max(0, indice - args.janela_contexto): min(len(linhas), indice + args.janela_contexto + 1)])

            if origem in {"biblioteca_terceiros", "codigo_gerado"}:
                classificacao = "ocorrencia_descartada"

            if eh_pipeline:
                if padrao.natureza_intervencao == "referencia_interface":
                    classificacao = "ocorrencia_descartada"
                elif not RE_EVIDENCIA_FORTE_PIPELINE.search(bloco_contexto):
                    classificacao = "ocorrencia_descartada"
                elif RE_METADATA_PIPELINE.search(linha) and not RE_EVIDENCIA_FORTE_PIPELINE.search(linha):
                    classificacao = "ocorrencia_descartada"

            if eh_config and RE_METADATA_PIPELINE.search(linha) and padrao.natureza_intervencao in {"referencia_interface", "contrato_integracao"}:
                classificacao = "ocorrencia_descartada"

            regiao = indice // max(args.janela_contexto, 1)
            if classificacao == "ocorrencia_descartada":
                prioridade = "descartada"
            elif classificacao == "adequacao_potencial":
                prioridade = "media"
            elif classificacao == "evidencia_contextual":
                prioridade = "baixa"
            elif classificacao == "adequacao_obrigatoria":
                prioridade = "critica" if padrao.pontuacao >= 95 else "alta"
            else:
                prioridade = padrao.prioridade_tratamento
            chave_dedup = (
                arquivo_rel,
                padrao.padrao_identificado,
                padrao.natureza_intervencao,
                classificacao,
                regiao,
            )
            if chave_dedup in vistos:
                continue
            vistos.add(chave_dedup)

            ocorrencias.append(
                OcorrenciaImpacto(
                    projeto=detectar_projeto(caminho, raiz, args.modo_agrupamento_projeto),
                    arquivo=arquivo_rel,
                    arquivo_absoluto=str(caminho.resolve()),
                    extensao=caminho.suffix.lower(),
                    camada=inferir_camada(caminho),
                    linha_inicial=indice + 1,
                    padrao_identificado=padrao.padrao_identificado,
                    categoria_tecnica=padrao.categoria_tecnica,
                    trecho=linha.strip()[:220],
                    contexto=" | ".join(linhas[max(0, indice - 1): min(len(linhas), indice + 2)])[:500],
                    pontuacao=padrao.pontuacao,
                    origem_artefato=origem,
                    artefato_gerado="sim" if origem in {"codigo_gerado", "snapshot_modelo", "arquivo_designer"} else "nao",
                    artefato_terceiro="sim" if origem == "biblioteca_terceiros" else "nao",
                    evidencia_contextual="sim" if contexto_ok else "nao",
                    prioridade_tratamento=prioridade,
                    identificador_consolidacao=gerar_identificador_consolidacao(
                        arquivo_rel,
                        regiao,
                        padrao.natureza_intervencao,
                        classificacao,
                        padrao.padrao_identificado,
                    ),
                    natureza_intervencao=padrao.natureza_intervencao,
                    classificacao_resultado=classificacao,
                )
            )
    return ocorrencias


def _coluna_excel(indice: int) -> str:
    letras = []
    while indice > 0:
        indice, resto = divmod(indice - 1, 26)
        letras.append(chr(65 + resto))
    return "".join(reversed(letras))


def _estilos_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="2">
    <font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/><family val="2"/></font>
  </fonts>
  <fills count="7">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF1F2937"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFECACA"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFEF3C7"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFDBEAFE"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFE5E7EB"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="10">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="3" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="5" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="6" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="3" borderId="0" xfId="0" applyFill="1" applyAlignment="1"><alignment wrapText="1" vertical="top"/></xf>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="0" xfId="0" applyFill="1" applyAlignment="1"><alignment wrapText="1" vertical="top"/></xf>
    <xf numFmtId="0" fontId="0" fillId="5" borderId="0" xfId="0" applyFill="1" applyAlignment="1"><alignment wrapText="1" vertical="top"/></xf>
    <xf numFmtId="0" fontId="0" fillId="6" borderId="0" xfId="0" applyFill="1" applyAlignment="1"><alignment wrapText="1" vertical="top"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""


def _gravar_xlsx_manual(caminho_xlsx: Path, ocorrencias: list[OcorrenciaImpacto]) -> None:
    estilo_base = {
        "adequacao_obrigatoria": 2,
        "adequacao_potencial": 3,
        "evidencia_contextual": 4,
        "ocorrencia_descartada": 5,
    }
    estilo_wrap = {
        "adequacao_obrigatoria": 6,
        "adequacao_potencial": 7,
        "evidencia_contextual": 8,
        "ocorrencia_descartada": 9,
    }

    col_trecho = COLUNAS_SAIDA.index("trecho") + 1
    col_contexto = COLUNAS_SAIDA.index("contexto") + 1
    col_classificacao = COLUNAS_SAIDA.index("classificacao_resultado")

    larguras = {i: max(12, len(nome) + 2) for i, nome in enumerate(COLUNAS_SAIDA, start=1)}
    linhas_xml = []

    # Cabeçalho
    celulas_cab = []
    for i, nome in enumerate(COLUNAS_SAIDA, start=1):
        ref = f"{_coluna_excel(i)}1"
        celulas_cab.append(f'<c r="{ref}" s="1" t="inlineStr"><is><t>{xml_escape(nome)}</t></is></c>')
    linhas_xml.append(f'<row r="1" ht="24" customHeight="1">{"".join(celulas_cab)}</row>')

    # Dados
    for idx, ocorrencia in enumerate(ocorrencias, start=2):
        valores = ocorrencia.para_linha()
        classificacao = str(valores[col_classificacao])
        celulas = []
        for col_idx, valor in enumerate(valores, start=1):
            texto = "" if valor is None else str(valor)
            larguras[col_idx] = min(80, max(larguras[col_idx], len(texto[:120]) + 2))
            ref = f"{_coluna_excel(col_idx)}{idx}"
            estilo = estilo_wrap.get(classificacao, 9) if col_idx in {col_trecho, col_contexto} else estilo_base.get(classificacao, 5)
            celulas.append(f'<c r="{ref}" s="{estilo}" t="inlineStr"><is><t>{xml_escape(texto)}</t></is></c>')
        linhas_xml.append(f'<row r="{idx}">{"".join(celulas)}</row>')

    cols_xml = []
    for idx in range(1, len(COLUNAS_SAIDA) + 1):
        cols_xml.append(f'<col min="{idx}" max="{idx}" width="{larguras[idx]:.2f}" customWidth="1"/>')

    ultima_coluna = _coluna_excel(len(COLUNAS_SAIDA))
    ultima_linha = max(1, len(ocorrencias) + 1)
    faixa_filtro = f"A1:{ultima_coluna}{ultima_linha}"

    sheet_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetViews>
    <sheetView workbookViewId="0">
      <pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>
    </sheetView>
  </sheetViews>
  <cols>{''.join(cols_xml)}</cols>
  <sheetData>{''.join(linhas_xml)}</sheetData>
  <autoFilter ref="{faixa_filtro}"/>
</worksheet>"""

    with zipfile.ZipFile(caminho_xlsx, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>""")
        zf.writestr("_rels/.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""")
        zf.writestr("xl/workbook.xml", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="levantamento_impacto" sheetId="1" r:id="rId1"/></sheets>
</workbook>""")
        zf.writestr("xl/_rels/workbook.xml.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>""")
        zf.writestr("xl/styles.xml", _estilos_xml())
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def gravar_xlsx(caminho_xlsx: Path, ocorrencias: list[OcorrenciaImpacto]) -> None:
    if not OPENPYXL_DISPONIVEL:
        _gravar_xlsx_manual(caminho_xlsx, ocorrencias)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "levantamento_impacto"
    ws.append(COLUNAS_SAIDA)

    preenchimentos = {
        "adequacao_obrigatoria": PatternFill("solid", fgColor="FFFECACA"),
        "adequacao_potencial": PatternFill("solid", fgColor="FFFEF3C7"),
        "evidencia_contextual": PatternFill("solid", fgColor="FFDBEAFE"),
        "ocorrencia_descartada": PatternFill("solid", fgColor="FFE5E7EB"),
    }

    fonte_cabecalho = Font(bold=True, color="FFFFFFFF")
    fundo_cabecalho = PatternFill("solid", fgColor="FF1F2937")
    for coluna, nome in enumerate(COLUNAS_SAIDA, start=1):
        cel = ws.cell(row=1, column=coluna, value=nome)
        cel.font = fonte_cabecalho
        cel.fill = fundo_cabecalho
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    coluna_classificacao = COLUNAS_SAIDA.index("classificacao_resultado") + 1
    colunas_textuais = {COLUNAS_SAIDA.index("trecho") + 1, COLUNAS_SAIDA.index("contexto") + 1}

    for ocorrencia in ocorrencias:
        ws.append(ocorrencia.para_linha())
        linha = ws.max_row
        classificacao = str(ws.cell(row=linha, column=coluna_classificacao).value)
        preenchimento = preenchimentos.get(classificacao)
        if preenchimento:
            for coluna in range(1, len(COLUNAS_SAIDA) + 1):
                ws.cell(row=linha, column=coluna).fill = preenchimento
        for coluna in colunas_textuais:
            ws.cell(row=linha, column=coluna).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, len(COLUNAS_SAIDA) + 1):
        letra = _coluna_excel(col_idx)
        maior = len(COLUNAS_SAIDA[col_idx - 1]) + 2
        for linha in range(2, ws.max_row + 1):
            valor = ws.cell(row=linha, column=col_idx).value
            texto = "" if valor is None else str(valor)
            maior = max(maior, min(len(texto), 120) + 2)
        ws.column_dimensions[letra].width = min(max(maior, 12), 80)

    wb.save(caminho_xlsx)


def _contagem(ocorrencias: list[OcorrenciaImpacto], atributo: str) -> dict[str, int]:
    total: dict[str, int] = {}
    for ocorrencia in ocorrencias:
        chave = getattr(ocorrencia, atributo)
        total[chave] = total.get(chave, 0) + 1
    return dict(sorted(total.items(), key=lambda item: (-item[1], item[0])))


def gravar_html(caminho_html: Path, ocorrencias: list[OcorrenciaImpacto]) -> None:
    resumo_classificacao = _contagem(ocorrencias, "classificacao_resultado")
    resumo_natureza = _contagem(ocorrencias, "natureza_intervencao")
    resumo_origem = _contagem(ocorrencias, "origem_artefato")

    cabecalho = "".join(f"<th>{html.escape(coluna)}</th>" for coluna in COLUNAS_SAIDA)
    corpo = "".join(
        "<tr>" + "".join(f"<td>{html.escape(str(valor))}</td>" for valor in ocorrencia.para_linha()) + "</tr>"
        for ocorrencia in ocorrencias
    )

    def bloco_resumo(titulo: str, itens: dict[str, int]) -> str:
        linhas = "".join(f"<li><strong>{html.escape(chave)}</strong>: {valor}</li>" for chave, valor in itens.items())
        return f"<div class='bloco'><h3>{html.escape(titulo)}</h3><ul>{linhas}</ul></div>"

    documento = f"""<!doctype html>
<html lang='pt-BR'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Levantamento de Impacto — Adequação ao CNPJ Alfanumérico</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 20px; color: #1f2937; }}
h1 {{ margin-bottom: 8px; }}
.resumo {{ display: flex; gap: 16px; flex-wrap: wrap; margin: 16px 0; }}
.bloco {{ border: 1px solid #d1d5db; border-radius: 6px; padding: 10px 12px; min-width: 260px; }}
.bloco h3 {{ margin: 0 0 8px 0; font-size: 14px; }}
.bloco ul {{ margin: 0; padding-left: 18px; }}
table {{ border-collapse: collapse; width: 100%; font-size: 12px; }}
th, td {{ border: 1px solid #d1d5db; padding: 6px; vertical-align: top; }}
th {{ background: #f3f4f6; position: sticky; top: 0; }}
</style>
</head>
<body>
<h1>Levantamento de Impacto — Adequação ao CNPJ Alfanumérico</h1>
<p>Total de ocorrências: <strong>{len(ocorrencias)}</strong></p>
<div class='resumo'>
{bloco_resumo('Total por classificacao_resultado', resumo_classificacao)}
{bloco_resumo('Total por natureza_intervencao', resumo_natureza)}
{bloco_resumo('Total por origem_artefato', resumo_origem)}
</div>
<table>
<thead><tr>{cabecalho}</tr></thead>
<tbody>{corpo}</tbody>
</table>
</body>
</html>"""
    caminho_html.write_text(documento, encoding="utf-8")


def main() -> None:
    args = argumentos()
    raiz = Path(args.raiz).resolve()
    diretorio_saida = Path(args.diretorio_saida).resolve()

    extensoes = set(EXTENSOES_SUPORTADAS)
    if args.extensoes_adicionais:
        extensoes.update(ext if ext.startswith(".") else f".{ext}" for ext in args.extensoes_adicionais)

    diretorios_excluidos = set(DIRETORIOS_EXCLUIDOS_PADRAO)
    if args.diretorios_excluidos:
        diretorios_excluidos.update(d.lower() for d in args.diretorios_excluidos)

    print(f"Iniciando varredura em: {raiz}", flush=True)
    print(f"Janela de contexto: ±{args.janela_contexto}", flush=True)

    arquivos = list(iterar_arquivos(raiz, extensoes, diretorios_excluidos, args.tamanho_maximo_kb, diretorio_saida))
    print(f"Arquivos elegíveis: {len(arquivos)}", flush=True)

    ocorrencias: list[OcorrenciaImpacto] = []
    for indice, arquivo in enumerate(arquivos, start=1):
        ocorrencias.extend(varrer_arquivo(arquivo, raiz, args))
        if indice % 200 == 0 or indice == len(arquivos):
            print(f"Processados {indice}/{len(arquivos)} arquivos...", flush=True)

    ordem_prioridade = {
        "critica": 0,
        "alta": 1,
        "media": 2,
        "baixa": 3,
        "descartada": 4,
    }
    ordem_classificacao = {
        "adequacao_obrigatoria": 0,
        "adequacao_potencial": 1,
        "evidencia_contextual": 2,
        "ocorrencia_descartada": 3,
    }
    ocorrencias.sort(
        key=lambda o: (
            ordem_prioridade.get(o.prioridade_tratamento, 9),
            ordem_classificacao.get(o.classificacao_resultado, 9),
            -o.pontuacao,
            o.projeto,
            o.arquivo,
            o.linha_inicial,
        )
    )

    diretorio_saida.mkdir(parents=True, exist_ok=True)
    caminho_xlsx = diretorio_saida / "levantamento_impacto_cnpj.xlsx"
    caminho_html = diretorio_saida / "levantamento_impacto_cnpj.html"

    gravar_xlsx(caminho_xlsx, ocorrencias)
    if not args.desabilitar_html:
        gravar_html(caminho_html, ocorrencias)

    print(f"Ocorrências consolidadas: {len(ocorrencias)}", flush=True)
    print(f"XLSX: {caminho_xlsx}", flush=True)
    if args.desabilitar_html:
        print("HTML desabilitado por parâmetro.", flush=True)
    else:
        print(f"HTML: {caminho_html}", flush=True)


if __name__ == "__main__":
    main()
